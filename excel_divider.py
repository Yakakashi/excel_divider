from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import os
from openpyxl.styles import Font, Alignment

my_window = Tk()
my_window.title("Разделить файл")
my_window.geometry("300x200")

def split_file():
    try:
        ftypes = [('Excel файлы', '*.xlsx'), ('Все файлы', '*')]
        file_path = filedialog.askopenfilename(filetypes=ftypes)
        if file_path == "":
            messagebox.showinfo("", "Файл не выбран")
        elif file_path != "":
            workbook = openpyxl.open(file_path, read_only=True)
            sheet = workbook.active

            # задать диапазон
            last = sheet.max_row
            cells = sheet["A2":"D" + f"{last}"]

            # создание списка строк из файла excel
            list_rows = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                list_rows.append(row)

            # создание файлов с уникальными названиями
            list_filenames = []
            for author, name, year, quantity in cells:
                book = openpyxl.Workbook()
                new_sheet = book.active
                if not os.path.exists(f"{author.value}.xlsx"):
                    book.save(f"{author.value.strip()}.xlsx")
                if f"{author.value.strip()}.xlsx" not in list_filenames:
                    list_filenames.append(f"{author.value.strip()}.xlsx")

            # заполнение заголовков
            for el in list_filenames:
                new_book = openpyxl.open(el)
                sheet_new_book = new_book.active
                for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
                    for ind, element in enumerate(row):
                        sheet_new_book.cell(row=1, column=ind + 1).value = element.value
                        sheet_new_book.cell(row=1, column=ind + 1).font = Font(bold=True)
                        sheet_new_book.cell(row=1, column=ind + 1).alignment = Alignment(horizontal="left")
                        sheet_new_book.column_dimensions["A"].width = 24
                        sheet_new_book.column_dimensions["B"].width = 30
                        sheet_new_book.column_dimensions["C"].width = 13
                        sheet_new_book.column_dimensions["D"].width = 15
                new_book.save(el)
                new_book.close()

            # заполнение ячеек
            row_ind = 2
            for el in list_filenames:
                for item in list_rows:
                    if el.find(item[0].value.strip()) > -1:
                        new_book = openpyxl.open(el)
                        sheet_new_book = new_book.active
                        for i, elem in enumerate(item):
                            sheet_new_book.cell(row=row_ind, column=i + 1).value = elem.value
                            sheet_new_book.cell(row=row_ind, column=i + 1).alignment = Alignment(horizontal="left")
                        new_book.save(el)
                        new_book.close()
                        row_ind += 1
                row_ind = 2
            messagebox.showinfo("","Программа выполнена")
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showwarning("Предупреждение!","Выберите файл одного из следующих форматов: xlsx/xlsm/xltx/xltm")

open_button = Button(text="Выбрать файл", command=split_file, height=2)
open_button.config(background="#a0cf74")
open_button.grid(row=2, padx=100,pady=75)

my_window.mainloop()
